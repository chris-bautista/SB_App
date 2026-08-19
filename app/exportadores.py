"""
Generación de archivos descargables (Excel y SQL) a partir de Notas.

Vive aparte de routers/admin.py para que las rutas del router se queden
delgadas: reciben la petición, consultan la base y delegan aquí la
construcción del archivo.
"""

import io
from datetime import datetime

from fastapi import Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from app.models import Nota

MEDIA_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ENCABEZADO_CONCEPTOS = ["Tipo", "Descripción", "Posición", "Lado", "Cantidad", "Importe", "Subtotal"]


def _vehiculo(n: Nota) -> str:
    return f"{n.vehiculo_marca} {n.vehiculo_modelo} {n.vehiculo_anio or ''}".strip()


def nota_a_excel(nota: Nota) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nota"

    ws.append(["Auto Servicio Bautista — Nota de servicio"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Fecha", nota.creado_en.strftime("%d/%m/%Y")])
    ws.append(["Cliente", nota.cliente_nombre])
    ws.append(["Teléfono", nota.cliente_telefono or ""])
    ws.append(["Vehículo", _vehiculo(nota)])
    ws.append(["Placas", nota.vehiculo_placas or ""])
    ws.append(["Trabajo", nota.trabajo])
    ws.append([])

    fila_encabezado = ws.max_row + 1
    ws.append(ENCABEZADO_CONCEPTOS)
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True)
    for c in nota.conceptos:
        ws.append([c.tipo, c.descripcion, c.posicion, c.lado, c.cantidad, c.importe, c.cantidad * c.importe])

    ws.append([])
    ws.append(["", "", "", "", "", "Total", nota.total])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True)

    for col, ancho in zip("ABCDEFG", (14, 30, 12, 12, 10, 12, 12)):
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer, media_type=MEDIA_XLSX,
        headers={"Content-Disposition": f'attachment; filename="nota_{nota.id}.xlsx"'},
    )


def notas_a_excel(notas: list[Nota]) -> StreamingResponse:
    wb = Workbook()

    ws_notas = wb.active
    ws_notas.title = "Notas"
    ws_notas.append(["ID", "Fecha", "Cliente", "Teléfono", "Vehículo", "Placas", "Trabajo", "Total"])
    for celda in ws_notas[1]:
        celda.font = Font(bold=True)
    for n in notas:
        ws_notas.append([
            n.id, n.creado_en.strftime("%d/%m/%Y"), n.cliente_nombre, n.cliente_telefono or "",
            _vehiculo(n), n.vehiculo_placas or "", n.trabajo, n.total,
        ])
    for col, ancho in zip("ABCDEFGH", (6, 12, 24, 14, 26, 12, 30, 12)):
        ws_notas.column_dimensions[col].width = ancho

    ws_conceptos = wb.create_sheet("Conceptos")
    ws_conceptos.append(["Nota ID", *ENCABEZADO_CONCEPTOS])
    for celda in ws_conceptos[1]:
        celda.font = Font(bold=True)
    for n in notas:
        for c in n.conceptos:
            ws_conceptos.append([
                n.id, c.tipo, c.descripcion, c.posicion, c.lado, c.cantidad, c.importe, c.cantidad * c.importe,
            ])
    for col, ancho in zip("ABCDEFGH", (8, 12, 30, 12, 12, 10, 12, 12)):
        ws_conceptos.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer, media_type=MEDIA_XLSX,
        headers={"Content-Disposition": 'attachment; filename="notas_historico.xlsx"'},
    )


def _valor_sql(valor) -> str:
    if valor is None:
        return "NULL"
    if isinstance(valor, bool):
        return "1" if valor else "0"
    if isinstance(valor, (int, float)):
        return str(valor)
    if isinstance(valor, datetime):
        return "'" + valor.strftime("%Y-%m-%d %H:%M:%S") + "'"
    return "'" + str(valor).replace("'", "''") + "'"


def notas_a_sql(notas: list[Nota]) -> Response:
    columnas_nota = [
        "id", "creado_en", "cliente_nombre", "cliente_telefono",
        "vehiculo_marca", "vehiculo_modelo", "vehiculo_anio", "vehiculo_placas",
        "trabajo", "aviso",
    ]
    columnas_concepto = ["id", "nota_id", "tipo", "descripcion", "posicion", "lado", "cantidad", "importe"]

    lineas = [
        "-- Exportado desde el panel de administración de Auto Servicio Bautista",
        f"-- Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
        "",
    ]
    for n in notas:
        valores = [getattr(n, col) for col in columnas_nota]
        lineas.append(
            f"INSERT INTO notas ({', '.join(columnas_nota)}) "
            f"VALUES ({', '.join(_valor_sql(v) for v in valores)});"
        )
        for c in n.conceptos:
            valores_c = [getattr(c, col) for col in columnas_concepto]
            lineas.append(
                f"INSERT INTO conceptos ({', '.join(columnas_concepto)}) "
                f"VALUES ({', '.join(_valor_sql(v) for v in valores_c)});"
            )

    contenido = "\n".join(lineas) + "\n"
    return Response(
        content=contenido, media_type="application/sql",
        headers={"Content-Disposition": 'attachment; filename="notas_historico.sql"'},
    )
